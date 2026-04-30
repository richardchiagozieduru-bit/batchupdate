import os
import uuid
import json
import logging
import pandas as pd
from io import BytesIO
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse, FileResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.views.decorators.cache import never_cache
from django.views.decorators.http import require_http_methods, require_POST
from django_q.tasks import async_task

from django.contrib.auth.models import User
from django.db import transaction

from openpyxl import load_workbook
from .models import UploadSession, ColumnMapping, MappingTemplate, Subscriber
from .services import (
    clean_dataframe,
    calculate_file_hash, read_uploaded_file,
    MAX_EXCEL_FILE_SIZE_MB, MAX_CSV_FILE_SIZE_MB,
    get_excel_sheet_names, read_uploaded_file_sheet,
    generate_sql_script, upload_raw_to_batchupdate,
    detect_header_row, build_sheet_name, get_subscribers_from_batchupdate,
    extract_sub_id,
)
from acctmgt.utils import is_external as _is_external, require_bound as _require_bound

logger = logging.getLogger(__name__)



@login_required
def upload_view(request):
    """Excel file upload page — handles multi-sheet files by splitting into one session per sheet."""
    # Gate: external users must be bound before they can upload
    if _require_bound(request.user):
        return redirect('redeem_token')

    # Resolve subscriber for this user
    if _is_external(request.user):
        subscriber = request.user.subscriber_profile.subscriber
        subscribers = None  # external users have no dropdown
    else:
        subscriber = None
        subscribers = get_subscribers_from_batchupdate()  # live from BatchUpdate Sheet1

    if request.method == 'POST' and request.FILES.get('excel_file'):
        assign_subscriber = request.POST.get('assign_subscriber', 'on')

        if assign_subscriber == 'on':
            # Internal users pick a subscriber per upload
            if not _is_external(request.user):
                sub_id = request.POST.get('subscriber', '').strip()
                try:
                    sub_id_int = int(float(sub_id))
                except (ValueError, TypeError):
                    messages.error(request, 'Please select a valid subscriber.')
                    return redirect('upload')
                sub_list = subscribers  # reuse already-fetched list
                sub_match = next((s for s in sub_list if s['subscriber_id'] == sub_id_int), None)
                if not sub_match:
                    messages.error(request, 'Please select a valid subscriber.')
                    return redirect('upload')
                subscriber, _ = Subscriber.objects.get_or_create(
                    subscriber_id=sub_id_int,
                    defaults={'subscriber_name': sub_match['subscriber_name']},
                )

            if not subscriber:
                messages.error(request, 'No subscriber assigned. Please contact admin.')
                return redirect('upload')
            excel_file = request.FILES['excel_file']

            # Validate file extension
            if not excel_file.name.endswith(('.xlsx', '.xls', '.csv')):
                messages.error(request, 'Please upload an Excel or CSV file (.xlsx, .xls, .csv)')
                return redirect('upload')

            # P1: Type-aware file size limit
            # Excel: openpyxl loads ~6x the file size into RAM, so a lower cap is enforced.
            # CSV: streamed in chunks, so a higher cap is safe.
            is_csv = excel_file.name.endswith('.csv')
            max_size_mb = MAX_CSV_FILE_SIZE_MB if is_csv else MAX_EXCEL_FILE_SIZE_MB
            if excel_file.size > max_size_mb * 1024 * 1024:
                messages.error(request, f'File too large. Maximum size is {max_size_mb} MB for {"CSV" if is_csv else "Excel"} files.')
                return redirect('upload')
            
            # Save the uploaded file temporarily to read sheets
            temp_session = UploadSession.objects.create(
                user=request.user,
                original_file=excel_file,
                original_filename=excel_file.name,
                status='pending_mapping',
                subscriber=subscriber,
            )
            file_path = temp_session.original_file.path
            
            # P2: Duplicate detection (check only recent sessions to avoid delay)
            try:
                file_hash = calculate_file_hash(file_path)
                existing = UploadSession.objects.filter(
                    user=request.user,
                    status='uploaded'
                ).exclude(id=temp_session.id).order_by('-uploaded_at')[:5]
                
                for existing_session in existing:
                    if existing_session.original_file and os.path.exists(existing_session.original_file.path):
                        existing_hash = calculate_file_hash(existing_session.original_file.path)
                        if file_hash == existing_hash:
                            messages.warning(
                                request,
                                f'This file appears to be a duplicate of "{existing_session.original_filename}" uploaded on {existing_session.uploaded_at.strftime("%Y-%m-%d")}. Proceeding anyway.'
                            )
                            break
            except Exception:
                logger.warning("Duplicate detection failed", exc_info=True)
            
            # Detect sheets — clean up temp session if the file cannot be read
            try:
                sheet_names = get_excel_sheet_names(file_path)
            except Exception as exc:
                logger.error(f"Could not read sheets from uploaded file: {exc}", exc_info=True)
                try:
                    temp_session.original_file.delete(save=False)
                    temp_session.delete()
                except Exception:
                    pass
                messages.error(request, 'Could not read the uploaded file. Please ensure it is a valid Excel or CSV file.')
                return redirect('upload')
            
            if sheet_names is None or len(sheet_names) <= 1:
                # Single sheet or CSV — use the temp session directly
                sheet_name = build_sheet_name(subscriber)
                temp_session.sheet_name = sheet_name
                temp_session.save()

                # Generate SQL script and upload to BatchUpdate
                _process_sheet_upload(temp_session, file_path, sheet_name)
                
                # Detect header row then read with it
                first_session = temp_session
                try:
                    template_signatures = list(
                        MappingTemplate.objects.filter(user=request.user)
                        .values_list('header_signature', flat=True)
                    )
                    hrow = detect_header_row(file_path, template_signatures=template_signatures)
                    first_session.header_row = hrow
                    first_session.save()
                    df = read_uploaded_file(file_path, header=hrow)
                    _try_auto_map(request, first_session, df)
                    if first_session.mappings.exists():
                        return redirect('process', session_id=first_session.id)
                except Exception:
                    logger.warning(f"Header detection / auto-map failed for session {first_session.id}", exc_info=True)
                
                return redirect('mapping', session_id=first_session.id)
            
            else:
                # Multi-sheet Excel — split into one session per sheet
                # Delete the temp session, we'll create individual ones
                temp_session.delete()

                batch_id = uuid.uuid4()
                created_sessions = []
                template_signatures = list(
                    MappingTemplate.objects.filter(user=request.user)
                    .values_list('header_signature', flat=True)
                )
                for sheet_index, sname in enumerate(sheet_names, start=1):
                    try:
                        hrow = detect_header_row(file_path, sheet_name=sname, template_signatures=template_signatures)
                        df = read_uploaded_file_sheet(file_path, sheet_name=sname, header=hrow)
                    except Exception as e:
                        messages.warning(request, f'Could not read sheet "{sname}": {e}')
                        continue

                    sheet_name = build_sheet_name(subscriber, index=sheet_index if sheet_index > 1 else None)
                    sheet_filename = f"{sheet_name}.xlsx"
                    sheet_dir = os.path.join('media', 'uploads')
                    os.makedirs(sheet_dir, exist_ok=True)
                    sheet_path = os.path.join(sheet_dir, sheet_filename)
                    df.to_excel(sheet_path, index=False, engine='openpyxl')

                    # df was read with header=hrow, so when saved via to_excel the
                    # resulting file always has headers at row 0 (pre-header rows are dropped).
                    session = UploadSession.objects.create(
                        user=request.user,
                        original_file=f'uploads/{sheet_filename}',
                        original_filename=sname,
                        status='pending_mapping',
                        sheet_name=sheet_name,
                        batch_id=batch_id,
                        source_filename=excel_file.name,
                        header_row=0,
                        subscriber=subscriber,
                    )

                    # Generate SQL script and upload to BatchUpdate
                    _process_sheet_upload(session, sheet_path, sheet_name, df=df)

                    # Auto-map immediately while df is in scope, then release it.
                    # Do NOT accumulate (session, df) tuples — each sheet's df can be
                    # hundreds of MB; holding all sheets simultaneously exhausts RAM.
                    try:
                        _try_auto_map(request, session, df)
                        if session.mappings.filter(target_column__isnull=False).exclude(target_column='').exists():
                            session.status = 'processing'
                            session.save()
                            async_task('update.tasks.process_file_task', session.id)
                    except Exception:
                        logger.warning(f"Auto-map failed for session {session.id}", exc_info=True)

                    created_sessions.append(session)  # only the session; df goes out of scope here

                if not created_sessions:
                    messages.error(request, 'No sheets could be read from the file.')
                    return redirect('upload')

                return redirect('batch', batch_id=batch_id)

        else:
            # Free upload mode (admin only) — subscriber resolved from filename/sheet tab name
            excel_files = request.FILES.getlist('excel_file')
            template_signatures = list(
                MappingTemplate.objects.filter(user=request.user)
                .values_list('header_signature', flat=True)
            )
            return _handle_free_upload(request, excel_files, template_signatures)
    
    # Group batch uploads: show one row per batch_id, individual rows for single uploads.
    # Cap at 10 entries.
    _all = UploadSession.objects.filter(
        user=request.user
    ).select_related('subscriber').order_by('-uploaded_at')[:60]
    seen_batches = set()
    sessions = []
    for s in _all:
        if s.batch_id:
            if s.batch_id not in seen_batches:
                seen_batches.add(s.batch_id)
                sessions.append(s)
        else:
            sessions.append(s)
        if len(sessions) >= 10:
            break

    templates = MappingTemplate.objects.filter(user=request.user)[:5]
    return render(request, 'update/upload.html', {
        'sessions': sessions,
        'templates': templates,
        'subscribers': subscribers,
        'subscriber': subscriber,
        'is_external': _is_external(request.user),
    })


def _process_sheet_upload(session, file_path, sheet_name, df=None):
    """Generate SQL script at upload time. BatchUpdate upload happens after cleaning in tasks.py."""
    import logging
    logger = logging.getLogger(__name__)

    # Generate SQL script only — BatchUpdate upload deferred until after cleaning
    try:
        script_path = generate_sql_script(sheet_name)
        session.generated_script = os.path.relpath(script_path, 'media')
        session.save()
    except Exception as e:
        logger.error(f"SQL script generation failed for sheet '{sheet_name}': {e}", exc_info=True)


def _resolve_subscriber_from_name(name_str):
    parts = name_str.split('_')
    try:
        sub_id_int = int(parts[0])
    except (ValueError, IndexError):
        raise ValueError(f'Could not parse subscriber ID from "{name_str}"')
    # Name is everything after the date segment (index 2 onward)
    sub_name = '_'.join(parts[2:]) if len(parts) >= 3 else name_str
    return sub_id_int, sub_name


def _handle_free_upload(request, excel_files, template_signatures):
    """
    Free upload mode (toggle OFF, admin only).
    Multiple files  → multi-excel mode: each filename drives its own subscriber.
    Single file     → multisheet mode: each sheet tab name drives its own subscriber.
    """
    batch_id = uuid.uuid4()
    created_sessions = []
    # Free upload mode only accepts Excel (multi-file) or Excel/CSV (single file)
    max_excel_size = MAX_EXCEL_FILE_SIZE_MB * 1024 * 1024
    max_csv_size = MAX_CSV_FILE_SIZE_MB * 1024 * 1024

    if len(excel_files) > 1:
        # ── Multi-Excel mode ──────────────────────────────────────────────────
        for f in excel_files:
            sheet_name = os.path.splitext(f.name)[0]
            try:
                sub_id_int, sub_name = _resolve_subscriber_from_name(sheet_name)
            except ValueError:
                messages.warning(request, f'Skipping "{f.name}": filename does not match expected pattern (subid_ddmmyyyy_name).')
                continue

            if f.size > max_excel_size:
                messages.warning(request, f'Skipping "{f.name}": file too large (max {MAX_EXCEL_FILE_SIZE_MB} MB).')
                continue
            if not f.name.endswith(('.xlsx', '.xls')):
                messages.warning(request, f'Skipping "{f.name}": must be .xlsx or .xls in multi-file mode.')
                continue

            subscriber, _ = Subscriber.objects.get_or_create(
                subscriber_id=sub_id_int,
                defaults={'subscriber_name': sub_name},
            )

            session = UploadSession.objects.create(
                user=request.user,
                original_file=f,
                original_filename=f.name,
                status='pending_mapping',
                sheet_name=sheet_name,
                batch_id=batch_id,
                source_filename=f.name,
                subscriber=subscriber,
            )
            file_path = session.original_file.path

            # Duplicate detection
            try:
                file_hash = calculate_file_hash(file_path)
                for existing_session in UploadSession.objects.filter(user=request.user, status='uploaded').exclude(id=session.id).order_by('-uploaded_at')[:5]:
                    if existing_session.original_file and os.path.exists(existing_session.original_file.path):
                        if file_hash == calculate_file_hash(existing_session.original_file.path):
                            messages.warning(request, f'"{f.name}" appears to be a duplicate of "{existing_session.original_filename}". Proceeding anyway.')
                            break
            except Exception:
                logger.warning("Duplicate detection failed", exc_info=True)

            _process_sheet_upload(session, file_path, sheet_name)

            try:
                hrow = detect_header_row(file_path, template_signatures=template_signatures)
                session.header_row = hrow
                session.save()
                df = read_uploaded_file(file_path, header=hrow)
                _try_auto_map(request, session, df)
                if session.mappings.filter(target_column__isnull=False).exclude(target_column='').exists():
                    session.status = 'processing'
                    session.save()
                    async_task('update.tasks.process_file_task', session.id)
            except Exception:
                logger.warning(f"Header detection / auto-map failed for session {session.id}", exc_info=True)
            messages.error(request, 'No files could be processed. Check that filenames follow the pattern: subid_ddmmyyyy_name.xlsx')
            return redirect('upload')

        return redirect('batch', batch_id=batch_id)

    else:
        # ── Single-file multisheet mode ────────────────────────────────────────
        f = excel_files[0]

        if f.size > (max_csv_size if f.name.endswith('.csv') else max_excel_size):
            limit = MAX_CSV_FILE_SIZE_MB if f.name.endswith('.csv') else MAX_EXCEL_FILE_SIZE_MB
            messages.error(request, f'File too large. Maximum size is {limit} MB.')
            return redirect('upload')
        if not f.name.endswith(('.xlsx', '.xls', '.csv')):
            messages.error(request, 'Please upload an Excel or CSV file (.xlsx, .xls, .csv).')
            return redirect('upload')

        # Save temporarily to read sheet names
        temp_session = UploadSession.objects.create(
            user=request.user,
            original_file=f,
            original_filename=f.name,
            status='pending_mapping',
        )
        file_path = temp_session.original_file.path

        # Clean up temp session if the file cannot be read
        try:
            sheet_names = get_excel_sheet_names(file_path)
        except Exception as exc:
            logger.error(f"Could not read sheets from uploaded file: {exc}", exc_info=True)
            try:
                temp_session.original_file.delete(save=False)
                temp_session.delete()
            except Exception:
                pass
            messages.error(request, 'Could not read the uploaded file. Please ensure it is a valid Excel or CSV file.')
            return redirect('upload')

        if sheet_names is None or len(sheet_names) <= 1:
            # Single sheet — subscriber from filename
            sheet_name = os.path.splitext(f.name)[0]
            try:
                sub_id_int, sub_name = _resolve_subscriber_from_name(sheet_name)
            except ValueError:
                temp_session.delete()
                messages.error(request, 'Filename does not match the expected pattern (subid_ddmmyyyy_name).')
                return redirect('upload')

            subscriber, _ = Subscriber.objects.get_or_create(
                subscriber_id=sub_id_int,
                defaults={'subscriber_name': sub_name},
            )
            temp_session.sheet_name = sheet_name
            temp_session.subscriber = subscriber
            temp_session.save()

            # Rename the sheet tab inside the Excel file to match the filename convention
            if file_path.endswith(('.xlsx', '.xls')):
                try:
                    wb = load_workbook(file_path)
                    wb.active.title = sheet_name[:31]  # Excel max sheet name length
                    wb.save(file_path)
                except Exception:
                    logger.warning(f"Could not rename sheet tab in {file_path}", exc_info=True)

            _process_sheet_upload(temp_session, file_path, sheet_name)

            try:
                hrow = detect_header_row(file_path, template_signatures=template_signatures)
                temp_session.header_row = hrow
                temp_session.save()
                df = read_uploaded_file(file_path, header=hrow)
                _try_auto_map(request, temp_session, df)
                if temp_session.mappings.exists():
                    return redirect('process', session_id=temp_session.id)
            except Exception:
                logger.warning(f"Header detection / auto-map failed for session {temp_session.id}", exc_info=True)

            return redirect('mapping', session_id=temp_session.id)

        else:
            # Multisheet — subscriber from each sheet tab name
            temp_session.delete()

            for sname in sheet_names:
                try:
                    sub_id_int, sub_name = _resolve_subscriber_from_name(sname)
                except ValueError:
                    messages.warning(request, f'Skipping sheet "{sname}": tab name does not match expected pattern (subid_ddmmyyyy_name).')
                    continue

                try:
                    hrow = detect_header_row(file_path, sheet_name=sname, template_signatures=template_signatures)
                    df = read_uploaded_file_sheet(file_path, sheet_name=sname, header=hrow)
                except Exception as e:
                    messages.warning(request, f'Could not read sheet "{sname}": {e}')
                    continue

                subscriber, _ = Subscriber.objects.get_or_create(
                    subscriber_id=sub_id_int,
                    defaults={'subscriber_name': sub_name},
                )

                sheet_filename = f"{sname}.xlsx"
                sheet_dir = os.path.join('media', 'uploads')
                os.makedirs(sheet_dir, exist_ok=True)
                sheet_path = os.path.join(sheet_dir, sheet_filename)
                df.to_excel(sheet_path, index=False, engine='openpyxl')

                session = UploadSession.objects.create(
                    user=request.user,
                    original_file=f'uploads/{sheet_filename}',
                    original_filename=sname,
                    status='pending_mapping',
                    sheet_name=sname,
                    batch_id=batch_id,
                    source_filename=f.name,
                    header_row=0,
                    subscriber=subscriber,
                )

                _process_sheet_upload(session, sheet_path, sname, df=df)

                try:
                    _try_auto_map(request, session, df)
                    if session.mappings.filter(target_column__isnull=False).exclude(target_column='').exists():
                        session.status = 'processing'
                        session.save()
                        async_task('update.tasks.process_file_task', session.id)
                except Exception:
                    logger.warning(f"Auto-map failed for session {session.id}", exc_info=True)

                created_sessions.append(session)  # df goes out of scope here

            if not created_sessions:
                messages.error(request, 'No sheets could be processed. Check that sheet tab names follow the pattern: subid_ddmmyyyy_name')
                return redirect('upload')

            return redirect('batch', batch_id=batch_id)


def _try_auto_map(request, session, df):
    """Try to auto-apply a saved mapping template, with subscriber-based fallback."""
    headers = sorted(list(df.columns))
    header_signature = json.dumps(headers)

    # 1. Exact header match via saved MappingTemplate (fastest path)
    template = MappingTemplate.objects.filter(
        user=request.user,
        header_signature=header_signature
    ).first()

    if template:
        for header, target in template.mappings.items():
            ColumnMapping.objects.create(
                session=session,
                original_header=header,
                target_column=target
            )
        template.use_count += 1
        template.save()
        return

    # 2. Subscriber-based fallback: reuse mapping from a previous session for the same subscriber.
    #    This means re-uploads for the same subscriber never require manual re-mapping as long as
    #    the column structure is unchanged — even across different users or upload dates.
    if session.subscriber_id:
        prev_sessions = (
            UploadSession.objects
            .filter(subscriber_id=session.subscriber_id)
            .exclude(id=session.id)
            .prefetch_related('mappings')
            .order_by('-uploaded_at')[:10]
        )
        for prev in prev_sessions:
            prev_mapping_objs = list(prev.mappings.all())
            if not prev_mapping_objs:
                continue
            # Only consider sessions whose full header set matches the current file
            prev_all_headers = sorted(m.original_header for m in prev_mapping_objs)
            if prev_all_headers != headers:
                continue
            # Build a dict of only the columns that were actually mapped
            prev_map = {m.original_header: m.target_column for m in prev_mapping_objs if m.target_column}
            if not prev_map:
                continue
            # Apply the previous mappings to the new session
            for header, target in prev_map.items():
                ColumnMapping.objects.create(
                    session=session,
                    original_header=header,
                    target_column=target,
                )
            # Promote to a MappingTemplate so the next upload takes the fast path
            MappingTemplate.objects.update_or_create(
                user=request.user,
                header_signature=header_signature,
                defaults={
                    'name': f"Auto: {session.subscriber} ({session.original_filename[:25]})",
                    'mappings': prev_map,
                },
            )
            return


@login_required
def mapping_view(request, session_id):
    """Interactive column mapping page"""
    session = get_object_or_404(UploadSession, id=session_id, user=request.user)
    
    try:
        df = read_uploaded_file(session.original_file.path, header=session.header_row)
        headers = list(df.columns)
    except Exception as e:
        messages.error(request, f'Error reading file: {str(e)}')
        return redirect('upload')
    
    existing_mappings = {m.original_header: m.target_column for m in session.mappings.all()}
    target_columns = ColumnMapping.TARGET_COLUMNS
    
    if request.method == 'POST':
        session.mappings.all().delete()
        mappings_dict = {}
        
        for header in headers:
            target = request.POST.get(f'mapping_{header}', '')
            ColumnMapping.objects.create(
                session=session,
                original_header=header,
                target_column=target
            )
            if target:
                mappings_dict[header] = target
        
        # Always save/update the template when a subscriber is set (enables future auto-mapping).
        # Also honour the explicit "save template" checkbox for sessions without a subscriber.
        if mappings_dict and (session.subscriber_id or request.POST.get('save_template')):
            header_signature = json.dumps(sorted(headers))
            name = (
                f"Auto: {session.subscriber} ({session.original_filename[:25]})"
                if session.subscriber_id
                else f"Template from {session.original_filename[:30]}"
            )
            MappingTemplate.objects.update_or_create(
                user=request.user,
                header_signature=header_signature,
                defaults={
                    'name': name,
                    'mappings': mappings_dict,
                },
            )

        return redirect('process', session_id=session.id)
    
    return render(request, 'update/mapping.html', {
        'session': session,
        'headers': headers,
        'target_columns': target_columns,
        'existing_mappings': existing_mappings,
    })


@login_required
def process_view(request, session_id):
    """Start async processing and show progress page"""
    session = get_object_or_404(UploadSession, id=session_id, user=request.user)
    
    # Check if mappings exist
    mappings = session.mappings.filter(target_column__isnull=False).exclude(target_column='')
    if not mappings.exists():
        messages.error(request, 'No columns mapped. Please map at least one column.')
        return redirect('mapping', session_id=session.id)
    
    # Only start task if not already processing
    if session.status != 'processing':
        async_task('update.tasks.process_file_task', session_id)
    
    # Render processing page - polls session status
    return render(request, 'update/processing.html', {
        'session': session,
    })


@login_required
def task_progress_view(request, session_id):
    """API endpoint to check task progress by session status"""
    session = get_object_or_404(UploadSession, id=session_id, user=request.user)
    
    if session.status == 'processing':
        response = {'current': 50, 'total': 100, 'status': 'Processing...'}
    elif session.status == 'processed':
        response = {'current': 100, 'total': 100, 'status': 'Complete!', 'rows_processed': session.rows_processed}
    elif session.status == 'uploaded':
        response = {'current': 100, 'total': 100, 'status': 'Complete!', 'rows_processed': session.rows_processed, 'rows_uploaded': session.rows_uploaded}
    elif session.status == 'error':
        response = {'current': 100, 'total': 100, 'status': session.error_message, 'error': True}
    else:
        response = {'current': 0, 'total': 100, 'status': 'Waiting to start...'}
    
    return JsonResponse(response)


@login_required
def result_view(request, session_id):
    """Show processing results"""
    session = get_object_or_404(
        UploadSession.objects.select_related('subscriber'),
        id=session_id, user=request.user
    )
    mappings = session.mappings.filter(target_column__isnull=False).exclude(target_column='')
    external = _is_external(request.user)

    return render(request, 'update/result.html', {
        'session': session,
        'mappings': mappings,
        'is_external': external,
    })


@login_required
def download_view(request, session_id):
    """Download processed Excel file — blocked for external users."""
    if _is_external(request.user):
        messages.error(request, 'You do not have permission to download processed files.')
        return redirect('result', session_id=session_id)
    session = get_object_or_404(UploadSession, id=session_id, user=request.user)
    
    if not session.processed_file:
        messages.error(request, 'No processed file available')
        return redirect('result', session_id=session.id)
    
    file_path = session.processed_file.path
    base_name = session.sheet_name if session.sheet_name else os.path.splitext(session.original_filename)[0]

    fh = open(file_path, 'rb')
    response = FileResponse(
        fh,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="cleaned_{base_name}.xlsx"'
    response['Content-Length'] = os.path.getsize(file_path)
    return response


@login_required
def download_rejected_view(request, session_id):
    """Download rejected rows Excel file"""
    session = get_object_or_404(UploadSession, id=session_id, user=request.user)
    
    if not session.rejected_file:
        messages.error(request, 'No rejected rows file available')
        return redirect('result', session_id=session.id)
    
    file_path = session.rejected_file.path
    base_name = session.sheet_name if session.sheet_name else os.path.splitext(session.original_filename)[0]

    fh = open(file_path, 'rb')
    response = FileResponse(
        fh,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="rejected_{base_name}.xlsx"'
    response['Content-Length'] = os.path.getsize(file_path)
    return response


@login_required
@require_POST
def undo_upload_view(request, session_id):
    """Undo/rollback — revert session status back to processed."""
    session = get_object_or_404(UploadSession, id=session_id, user=request.user)

    if session.status != 'uploaded':
        messages.error(request, 'Cannot undo - no records were uploaded from this session.')
        return redirect('result', session_id=session.id)

    try:
        session.status = 'processed'
        session.rows_uploaded = 0
        session.batchupdate_uploaded = False
        session.save()
        messages.success(request, 'Session reverted to processed state.')
    except Exception as e:
        messages.error(request, f'Undo failed: {str(e)}')

    return redirect('result', session_id=session.id)


@login_required
@require_POST
def delete_session_view(request, session_id):
    """Delete an upload session and its associated data"""
    session = get_object_or_404(UploadSession, id=session_id, user=request.user)

    try:
        filename = session.original_filename
        session.delete()
        messages.success(request, f'Deleted session: {filename}')
    except Exception as e:
        messages.error(request, f'Delete failed: {str(e)}')

    return redirect('upload')


@login_required
@require_POST
def delete_batch_view(request, batch_id):
    """Delete all sessions belonging to a batch."""
    sessions = UploadSession.objects.filter(user=request.user, batch_id=batch_id)
    count = sessions.count()
    if count == 0:
        messages.error(request, 'Batch not found.')
    else:
        sessions.delete()
        messages.success(request, f'Deleted batch ({count} session{"s" if count != 1 else ""}).')
    return redirect('upload')


@login_required
def batch_view(request, batch_id):
    """Landing page for a multi-sheet batch upload."""
    sessions = UploadSession.objects.filter(
        user=request.user,
        batch_id=batch_id
    ).order_by('uploaded_at')
    
    if not sessions.exists():
        messages.error(request, 'Batch not found.')
        return redirect('upload')
    
    source_filename = sessions.first().source_filename or 'Unknown file'
    scripts_count = sessions.filter(generated_script__isnull=False).exclude(generated_script='').count()
    bu_count = sessions.filter(batchupdate_uploaded=True).count()
    uploaded_count = sessions.filter(status='uploaded').count()
    processed_count = sessions.filter(processed_file__isnull=False).exclude(processed_file='').count()
    has_processing = sessions.filter(status='processing').exists()
    
    return render(request, 'update/batch.html', {
        'sessions': sessions,
        'batch_id': batch_id,
        'source_filename': source_filename,
        'scripts_count': scripts_count,
        'bu_count': bu_count,
        'uploaded_count': uploaded_count,
        'processed_count': processed_count,
        'has_processing': has_processing,
        'is_external': _is_external(request.user),
    })


@login_required
def batch_progress_view(request, batch_id):
    """AJAX endpoint returning live status for every session in the batch."""
    sessions = UploadSession.objects.filter(
        user=request.user,
        batch_id=batch_id,
    ).order_by('uploaded_at')

    sheets = []
    for s in sessions:
        sheets.append({
            'id': s.id,
            'name': s.sheet_name,
            'status': s.status,
            'status_display': s.get_status_display(),
            'rows_processed': s.rows_processed,
            'rows_uploaded': s.rows_uploaded,
            'rows_rejected': s.rows_rejected,
            'batchupdate_uploaded': s.batchupdate_uploaded,
            'error_message': s.error_message or '',
        })

    total = len(sheets)
    done = sum(1 for s in sheets if s['status'] in ('uploaded', 'processed'))
    errored = sum(1 for s in sheets if s['status'] == 'error')
    processing = sum(1 for s in sheets if s['status'] == 'processing')
    all_done = (done + errored) == total and processing == 0

    return JsonResponse({
        'sheets': sheets,
        'total': total,
        'done': done,
        'errored': errored,
        'processing': processing,
        'all_done': all_done,
    })


@login_required
def download_batch_combined(request, batch_id):
    """Download all cleaned sheets from a batch as a single Excel workbook."""
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from .services import format_excel_sheet, to_excel_safe_sheet_name
    import io

    sessions = UploadSession.objects.filter(
        user=request.user,
        batch_id=batch_id
    ).order_by('uploaded_at')

    if not sessions.exists():
        messages.error(request, 'Batch not found.')
        return redirect('upload')

    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)

    sheets_added = 0
    source_filename = sessions.first().source_filename or str(batch_id)

    for session in sessions:
        if not session.processed_file:
            continue
        file_path = session.processed_file.path
        if not os.path.exists(file_path):
            continue
        try:
            df = pd.read_excel(file_path, dtype=str, engine='openpyxl')
        except Exception:
            continue

        # Restore numeric columns to actual numbers so General-formatted cells
        # are right-aligned (matching individual download behaviour).
        # Text columns (AccountNo, LoanClassification, AccountStatusCode) stay as strings.
        _numeric_cols = [
            'CurrentBalanceAmt', 'AmountOverdue', 'MonthsInArrears',
            'overdue_amount', 'months_in_arrears',
        ]
        for _col in _numeric_cols:
            if _col in df.columns:
                df[_col] = pd.to_numeric(df[_col], errors='coerce')

        sheet_title = to_excel_safe_sheet_name(session.sheet_name or session.original_filename)
        ws = wb.create_sheet(title=sheet_title)
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)
        format_excel_sheet(ws)
        sheets_added += 1

    if sheets_added == 0:
        messages.error(request, 'No processed sheets available yet. Complete mapping and processing first.')
        return redirect('batch', batch_id=batch_id)

    base_name = os.path.splitext(source_filename)[0]
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="cleaned_{base_name}.xlsx"'
    return response


@login_required
def download_script_view(request, session_id):
    """Download the generated SQL script for a session."""
    session = get_object_or_404(UploadSession, id=session_id, user=request.user)
    
    if not session.generated_script:
        messages.error(request, 'No SQL script available for this session.')
        return redirect('result', session_id=session.id)
    
    file_path = session.generated_script.path
    script_name = session.sheet_name or os.path.splitext(session.original_filename)[0]

    with open(file_path, 'r', encoding='utf-8') as f:
        response = HttpResponse(f.read(), content_type='application/sql')
        response['Content-Disposition'] = f'attachment; filename="{script_name}.sql"'
        return response


@login_required
def batch_mapping_view(request, batch_id):
    """Unified mapping page for all sheets in a batch — map every sheet, then process all at once."""
    sessions = UploadSession.objects.filter(
        user=request.user,
        batch_id=batch_id,
    ).order_by('uploaded_at')

    if not sessions.exists():
        messages.error(request, 'Batch not found.')
        return redirect('upload')

    # Build per-sheet data: headers, existing mappings, status
    sheets_data = []
    for session in sessions:
        if session.status not in ('pending_mapping', 'error'):
            sheets_data.append({
                'session': session,
                'headers': [],
                'existing_mappings': {},
                'already_mapped': True,
            })
            continue

        try:
            df = read_uploaded_file(session.original_file.path)
            headers = list(df.columns)
        except Exception as e:
            sheets_data.append({
                'session': session,
                'headers': [],
                'existing_mappings': {},
                'error': str(e),
                'already_mapped': False,
            })
            continue

        existing_mappings = {m.original_header: m.target_column for m in session.mappings.all()}
        sheets_data.append({
            'session': session,
            'headers': headers,
            'existing_mappings': existing_mappings,
            'already_mapped': False,
        })

    target_columns = ColumnMapping.TARGET_COLUMNS

    if request.method == 'POST':
        sessions_to_process = []

        for sheet in sheets_data:
            session = sheet['session']
            if sheet.get('already_mapped') or sheet.get('error'):
                continue

            headers = sheet['headers']
            session.mappings.all().delete()
            mappings_dict = {}

            for header in headers:
                target = request.POST.get(f'sess_{session.id}_mapping_{header}', '')
                ColumnMapping.objects.create(
                    session=session,
                    original_header=header,
                    target_column=target,
                )
                if target:
                    mappings_dict[header] = target

            if mappings_dict:
                sessions_to_process.append(session)

                if request.POST.get('save_template'):
                    header_signature = json.dumps(sorted(headers))
                    MappingTemplate.objects.update_or_create(
                        user=request.user,
                        header_signature=header_signature,
                        defaults={
                            'name': f"Template from {session.original_filename[:30]}",
                            'mappings': mappings_dict,
                        },
                    )

        for session in sessions_to_process:
            session.status = 'processing'
            session.save()
            async_task('update.tasks.process_file_task', session.id)

        if sessions_to_process:
            messages.success(request, f'Processing started for {len(sessions_to_process)} sheet(s).')
        else:
            messages.warning(request, 'No sheets were mapped. Please map at least one column per sheet.')

        return redirect('batch', batch_id=batch_id)

    source_filename = sessions.first().source_filename or 'Unknown file'
    pending_count = sum(1 for s in sheets_data if not s.get('already_mapped') and not s.get('error'))
    done_count = sum(1 for s in sheets_data if s.get('already_mapped'))

    return render(request, 'update/batch_mapping.html', {
        'sheets_data': sheets_data,
        'target_columns': target_columns,
        'batch_id': batch_id,
        'source_filename': source_filename,
        'pending_count': pending_count,
        'done_count': done_count,
    })


@login_required
def download_batch_scripts_zip(request, batch_id):
    """Download all SQL scripts for a batch as a single ZIP file."""
    import zipfile
    import io as _io

    sessions = UploadSession.objects.filter(
        user=request.user,
        batch_id=batch_id
    ).order_by('uploaded_at')

    if not sessions.exists():
        messages.error(request, 'Batch not found.')
        return redirect('upload')

    source_filename = sessions.first().source_filename or str(batch_id)
    buffer = _io.BytesIO()
    files_added = 0

    with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        for session in sessions:
            if not session.generated_script:
                continue
            script_path = session.generated_script.path
            if not os.path.exists(script_path):
                continue
            arcname = f"{session.sheet_name or session.original_filename}.sql"
            zf.write(script_path, arcname=arcname)
            files_added += 1

    if files_added == 0:
        messages.error(request, 'No SQL scripts available yet for this batch.')
        return redirect('batch', batch_id=batch_id)

    base_name = os.path.splitext(source_filename)[0]
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="scripts_{base_name}.zip"'
    return response
