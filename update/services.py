import io
import os
import re
import hashlib
import logging
import warnings
import zipfile
import msoffcrypto
import pyodbc
import numpy as np
import pandas as pd
from datetime import datetime
from decimal import Decimal, InvalidOperation
from django.conf import settings
from django.db import transaction
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils.exceptions import InvalidFileException

from .models import UploadSession

logger = logging.getLogger(__name__)


def detect_header_row(file_path, sheet_name=None, template_signatures=None, max_scan_rows=15):
    """
    Detect which row (0-based) contains the column headers.

    Strategy:
      1. If template_signatures (list of JSON strings) provided, scan rows and
         return the first row whose sorted values match a known signature.
      2. Fall back to heuristic scoring:
         - all strings, no blanks, no duplicates  → header-like
         - next row contains at least one numeric  → data-like
         Row 0 gets a tiebreaker bonus so default behaviour is unchanged.

    Returns: int (0-based index to pass as pandas header= parameter).
    """
    import json as _json

    ext = os.path.splitext(file_path)[1].lower()

    def _is_numeric(v):
        try:
            float(v)
            return True
        except (ValueError, TypeError):
            return False

    with warnings.catch_warnings():
        warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
        try:
            kw = dict(header=None, nrows=max_scan_rows, dtype=str)
            if ext == '.csv':
                raw = pd.read_csv(file_path, **kw)
            elif ext == '.xlsx':
                if sheet_name:
                    kw['sheet_name'] = sheet_name
                raw = pd.read_excel(file_path, engine='openpyxl', **kw)
            elif ext == '.xls':
                if sheet_name:
                    kw['sheet_name'] = sheet_name
                raw = pd.read_excel(file_path, engine='xlrd', **kw)
            else:
                if sheet_name:
                    kw['sheet_name'] = sheet_name
                try:
                    raw = pd.read_excel(file_path, engine='openpyxl', **kw)
                except Exception:
                    raw = pd.read_excel(file_path, engine='xlrd', **kw)
        except Exception:
            return 0

    n_rows = len(raw)
    if n_rows == 0:
        return 0

    sig_set = set(template_signatures) if template_signatures else set()

    # ── Step 1: exact template-signature match ──
    for row_idx in range(min(n_rows, max_scan_rows)):
        candidate = raw.iloc[row_idx].astype(str).str.strip().tolist()
        candidate_clean = [v for v in candidate if v and v.lower() not in ('nan', 'none', '')]
        if not candidate_clean:
            continue
        sig = _json.dumps(sorted(candidate_clean))
        if sig in sig_set:
            logger.info(f"Header row detected via template match at row {row_idx}: {file_path}")
            return row_idx

    # ── Step 2: heuristic scoring ──
    best_row, best_score = 0, -1
    scan_limit = min(n_rows - 1, max_scan_rows - 1)

    for row_idx in range(scan_limit + 1):
        row_vals = raw.iloc[row_idx].astype(str).str.strip().tolist()
        non_empty = [v for v in row_vals if v and v.lower() not in ('nan', 'none', '')]
        if not non_empty:
            continue

        score = 0
        if len(non_empty) == len(row_vals):          # no blanks
            score += 2
        if len(set(non_empty)) == len(non_empty):    # no duplicates
            score += 2
        if all(not _is_numeric(v) for v in non_empty):  # all look like text
            score += 3
        if row_idx + 1 < n_rows:                     # next row has a numeric value
            nxt = raw.iloc[row_idx + 1].astype(str).str.strip().tolist()
            if any(_is_numeric(v) for v in nxt if v and v.lower() not in ('nan', 'none', '')):
                score += 3
        if row_idx == 0:                             # tiebreaker: prefer row 0
            score += 1

        if score > best_score:
            best_score, best_row = score, row_idx

    logger.info(f"Header row detected via heuristic: row {best_row} (score={best_score}) in {file_path}")
    return best_row


def read_excel_file(uploaded_file, filename: str, password: str | None = None) -> pd.ExcelFile:
    """Read an uploaded Excel stream into a pd.ExcelFile, with optional password decryption.

    Args:
        uploaded_file: A file-like object (e.g. Django InMemoryUploadedFile or BytesIO).
        filename:      Original filename, used to select the correct engine.
        password:      Optional workbook password for encrypted Excel files.

    Returns:
        pd.ExcelFile ready for sheet inspection or DataFrame extraction.

    Raises:
        ValueError: If the password is wrong or the file is corrupt.
    """
    uploaded_file.seek(0)
    source_stream = uploaded_file

    if password:
        try:
            decrypted = io.BytesIO()
            office_file = msoffcrypto.OfficeFile(uploaded_file)
            office_file.load_key(password=password)
            office_file.decrypt(decrypted)
            decrypted.seek(0)
            source_stream = decrypted
        except Exception as e:
            raise ValueError("Incorrect password or the file is not encrypted.") from e

    try:
        if filename.lower().endswith(".xls"):
            return pd.ExcelFile(source_stream, engine="xlrd")
        return pd.ExcelFile(source_stream, engine="openpyxl")
    except zipfile.BadZipFile as e:
        raise ValueError("The Excel file is invalid or corrupt (not a valid zip-based workbook).") from e


def read_uploaded_file(file_path, header=0, nrows=None):
    """Read CSV or Excel file into a DataFrame with engine auto-detection.

    Pass nrows=0 to read only the header row (useful for column inspection
    without loading all data into memory).
    """
    ext = os.path.splitext(file_path)[1].lower()
    logger.info(f"Reading file: {file_path} (extension: {ext}, header_row={header}, nrows={nrows})")

    with warnings.catch_warnings():
        warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
        if ext == '.csv':
            kw = dict(dtype=str, header=header)
            if nrows is not None:
                kw['nrows'] = nrows
            return pd.read_csv(file_path, **kw)
        elif ext == '.xlsx':
            kw = dict(dtype=str, engine='openpyxl', header=header)
            if nrows is not None:
                kw['nrows'] = nrows
            return pd.read_excel(file_path, **kw)
        elif ext == '.xls':
            kw = dict(dtype=str, engine='xlrd', header=header)
            if nrows is not None:
                kw['nrows'] = nrows
            return pd.read_excel(file_path, **kw)
        else:
            kw = dict(dtype=str, header=header)
            if nrows is not None:
                kw['nrows'] = nrows
            try:
                return pd.read_excel(file_path, engine='openpyxl', **kw)
            except Exception:
                logger.warning(f"openpyxl failed for {file_path}, trying xlrd")
                return pd.read_excel(file_path, engine='xlrd', **kw)


def _zero_mask_width(number_format):
    """
    Return the number of '0' placeholders in a zero-mask number format, or None.

    A zero-mask is a format built entirely of '0' digits plus separators (dashes,
    spaces, parentheses), e.g. '000000', '0000-000000', '(000) 000-0000'.
    These formats prove a fixed display width with mandatory leading zeros.

    Formats containing '#', '.', '%', 'E', or 'e' are NOT zero-masks and return None.
    """
    if not number_format or number_format in ('General', '@', ''):
        return None
    fmt = number_format
    fmt = re.sub(r'"[^"]*"', '', fmt)   # remove quoted literals
    fmt = re.sub(r'\\.', '', fmt)        # remove backslash-escaped chars
    if re.search(r'[#.%eE]', fmt):
        return None
    zeros = fmt.count('0')
    return zeros if zeros > 0 else None


def _reconstruct_account_value(cell_value, data_type, number_format):
    """
    Reconstruct a single account number string from an openpyxl cell's value and style.

    - Text cells (data_type 's'): returned as-is — leading zeros already present.
    - Numeric cells with a zero-mask format (e.g. '000000'): left-padded with zeros
      to the mask width.
    - Numeric cells without a zero-mask: returned as plain string — no guessing.
    - None values: returned as None.
    """
    if cell_value is None:
        return None
    if data_type == 's':
        return str(cell_value)
    if data_type == 'n' and number_format:
        width = _zero_mask_width(number_format)
        if width is not None:
            try:
                return str(int(float(cell_value))).zfill(width)
            except (ValueError, TypeError):
                pass
    return str(cell_value)


def read_account_column_styled(file_path, sheet_name, header_row, col_name):
    """
    Read one column from an .xlsx file with full cell-style awareness to recover
    leading zeros that pandas strips during value-only ingestion.

    For each data cell (below the header row):
      - Text cell (data_type='s'): kept as-is.
      - Numeric cell with a zero-mask format (e.g. '000000'): left-padded to mask width.
      - Numeric cell with no zero-mask: returned as plain string (no guessing).

    Returns a list of strings aligned to the data rows (header excluded), or None if
    the file is not .xlsx, the column is not found, or the workbook cannot be opened.

    Note: opens the workbook in full (non-read-only) mode to access cell styles.
    For large files this temporarily uses more RAM than the streaming path; this is
    the unavoidable cost of reading style metadata.
    """
    if os.path.splitext(file_path)[1].lower() != '.xlsx':
        return None
    with warnings.catch_warnings():
        warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
        try:
            wb = load_workbook(file_path, data_only=True)
        except Exception as exc:
            logger.warning(f"read_account_column_styled: could not open {file_path}: {exc}")
            return None
        try:
            ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
            header_row_1 = header_row + 1  # openpyxl rows are 1-based
            # Locate the target column by matching the header cell value
            col_idx = None
            for cell in ws[header_row_1]:
                if cell.value == col_name:
                    col_idx = cell.column - 1  # convert to 0-based
                    break
            if col_idx is None:
                logger.debug(f"read_account_column_styled: '{col_name}' not found in row {header_row_1}")
                return None
            result = []
            for row in ws.iter_rows(min_row=header_row_1 + 1):
                cell = row[col_idx] if col_idx < len(row) else None
                result.append(
                    _reconstruct_account_value(cell.value, cell.data_type, cell.number_format)
                    if cell is not None else None
                )
            logger.info(
                f"read_account_column_styled: '{col_name}' — {len(result)} rows read from {file_path}"
            )
            return result
        finally:
            wb.close()


# Configuration
# Memory note: openpyxl loads ~6x the file size into RAM when parsing xlsx.
# Two concurrent Django-Q workers at 50 MB each = ~600 MB RAM just for parsing.
# Excel files over LARGE_FILE_THRESHOLD_MB are stream-converted to a temp CSV
# before processing so peak RAM stays flat regardless of file size.
MAX_EXCEL_FILE_SIZE_MB = 200   # Hard cap for .xlsx / .xls uploads
MAX_CSV_FILE_SIZE_MB = 500     # Hard cap for .csv uploads
MAX_FILE_SIZE_MB = max(MAX_EXCEL_FILE_SIZE_MB, MAX_CSV_FILE_SIZE_MB)  # kept for any external references
LARGE_FILE_THRESHOLD_MB = 50   # xlsx files over this are converted to CSV before processing
CSV_CHUNK_SIZE = 50_000         # rows per chunk during CSV processing

# Target column metadata — single source of truth in update/columns.py
from .columns import (
    NUMERIC_COLUMNS, TEXT_COLUMNS, STRING_COLUMNS,
    VALIDATION_RULES, DISPLAY_HEADERS,
)

# Value transformation mappings
ACCOUNT_STATUS_MAP = {
    'Open': ['001', 'open', '01', '1', 'opened', 'active'],
    'Closed': ['002', 'closed', 'close', '02', '2', 'cloed'],
    'Writtenoff': ['003', 'written off', 'written0ff', '03', '3', 'writenoff', 'writtenoff', 'writeoff', 'write off', ',write off'],
    'Performing': ['performing', 'perfroming']
}

LOAN_CLASSIFICATION_MAP = {
    'Performing': ['001', 'performing', '1', '01', 'perform', 'performingloansperformingadvances', 
                   'performing loans', 'performing advances', 'performingloans', 'performingadvances', 'performimg'],
    'Watchlist': ['002', 'watchlist', '02', '2', 'pass and watch', 'passwatch', 'pass watch', 
                  'paasandwatch', 'pw', 'p&w', 'passandwatch'],
    'Sub standard': ['003', 'sub standard', 'substandard', '03', '3', 'sub', 'subs', 
                     'substandardloans', 'substandardadvances'],
    'Doubtful': ['004', 'doubtful', '04', '4', 'very doubtful', 'verydoubtful', 'doub', 'doubt', 
                 'doubtfulloans', 'doubtfuladvances'],
    'Lost': ['005', 'lost', '05', '5', 'loss', 'l'],
    'Writtenoff': ['write off', 'written off', 'writeoff']
}


def calculate_file_hash(file_path):
    """Calculate MD5 hash of a file for duplicate detection."""
    hasher = hashlib.md5()
    with open(file_path, 'rb') as f:
        for chunk in iter(lambda: f.read(8192), b''):
            hasher.update(chunk)
    return hasher.hexdigest()


def normalize_value(value, mapping_dict):
    """
    Normalize a value using a mapping dictionary.
    Converts variations to their canonical form.
    """
    if pd.isna(value) or value is None:
        return None
    
    value_str = str(value).lower().strip()
    
    for canonical, variations in mapping_dict.items():
        if value_str in [v.lower() for v in variations]:
            return canonical
    
    return value_str


def validate_numeric(value, column_name):
    """Validate numeric value against rules. Returns (valid, value)."""
    if value is None:
        return True, None
    
    rules = VALIDATION_RULES.get(column_name, {})
    min_val = rules.get('min')
    max_val = rules.get('max')
    
    try:
        num_val = float(value) if not isinstance(value, (int, float, Decimal)) else float(value)
        
        if min_val is not None and num_val < min_val:
            return False, None
        if max_val is not None and num_val > max_val:
            return False, None
            
        return True, value
    except (ValueError, TypeError):
        return False, None


def clean_value(value, column_name):
    """
    Clean a single value based on column type.
    - Numeric: Remove special chars except decimal point, validate
    - Text: Remove all special chars, convert to string
    - Special columns: Apply value transformations
    """
    try:
        if pd.isna(value):
            return None
    except (ValueError, TypeError):
        return None
    
    value_str = str(value).strip()
    
    if column_name in NUMERIC_COLUMNS:
        # Keep only digits and decimal point
        cleaned = re.sub(r'[^\d.]', '', value_str)
        if not cleaned:
            return None
        try:
            float_val = float(cleaned)
            # CurrentBalanceAmt and overdue_amount keep decimal precision
            # months_in_arrears is always a whole number
            if column_name in ('CurrentBalanceAmt', 'overdue_amount'):
                result = float_val
            else:
                result = int(float_val)

            # Validate against rules
            valid, validated = validate_numeric(result, column_name)
            return validated if valid else None

        except (ValueError, InvalidOperation):
            return None
    
    elif column_name == 'account_status_code':
        return normalize_value(value, ACCOUNT_STATUS_MAP)
    
    elif column_name == 'loan_classification':
        return normalize_value(value, LOAN_CLASSIFICATION_MAP)
    
    elif column_name in STRING_COLUMNS:
        return value_str if value_str else None
    
    else:
        cleaned = re.sub(r'[^a-zA-Z0-9\s]', '', value_str)
        return cleaned.strip() if cleaned.strip() else None


def is_row_empty(row, columns):
    """Check if all mapped columns in a row are empty/None."""
    for col in columns:
        val = row.get(col)
        if val is not None and str(val).strip():
            return False
    return True


def clean_dataframe(df, mappings, format_for_display=False):
    """
    Clean dataframe based on column mappings.
    Applies business validation rules, skips invalid rows, removes duplicates.
    
    Returns: (cleaned_df, rejected_df)
    """
    columns_to_keep = list(mappings.keys())
    df_subset = df[columns_to_keep].copy()
    df_subset.rename(columns=mappings, inplace=True)
    
    # Handle duplicate target columns (e.g. two source cols mapped to same target)
    # Keep only the first occurrence of each column name
    if df_subset.columns.duplicated().any():
        dupes = df_subset.columns[df_subset.columns.duplicated(keep=False)].unique().tolist()
        logger.warning(f"Duplicate target columns detected and deduplicated: {dupes}")
        df_subset = df_subset.loc[:, ~df_subset.columns.duplicated(keep='first')]
    
    # Clean each column
    for column in df_subset.columns:
        df_subset[column] = df_subset[column].apply(lambda x: clean_value(x, column))
    
    # Remove empty rows (where all mapped columns are None) - NOT tracked as rejected
    mapped_columns = list(mappings.values())
    df_subset = df_subset.dropna(subset=mapped_columns, how='all')
    
    # Ensure all target columns exist so auto-fill logic can inject defaults
    # for columns that were not mapped (e.g. overdue_amount missing from source file)
    _ALL_TARGET_COLS = ['account_number', 'CurrentBalanceAmt', 'overdue_amount',
                        'months_in_arrears', 'loan_classification', 'account_status_code']
    _injected = []
    for _col in _ALL_TARGET_COLS:
        if _col not in df_subset.columns:
            df_subset[_col] = None
            _injected.append(_col)
    if _injected:
        logger.info(f"Auto-injected missing columns (will be filled with 0 where balance=0): {_injected}")
    
    # --- Business Validation Rules (returns valid + rejected) ---
    # Pass all target columns (including injected ones) so auto-fill for balance=0 works
    all_target_cols = list(df_subset.columns)
    df_subset, rejected_df = apply_business_rules(df_subset, mapped_columns, all_target_cols)
    
    # Remove duplicate rows (where ALL values match exactly)
    df_subset = df_subset.drop_duplicates(keep='first')
    
    # Reset index after filtering
    df_subset = df_subset.reset_index(drop=True)
    
    # Rename columns for display (Proper Case without underscores)
    if format_for_display:
        df_subset.rename(columns=DISPLAY_HEADERS, inplace=True)
        # Also rename rejected df columns for display
        if not rejected_df.empty:
            reason_col = rejected_df['Rejection Reason']
            rejected_df = rejected_df.drop(columns=['Rejection Reason'])
            rejected_df.rename(columns=DISPLAY_HEADERS, inplace=True)
            rejected_df['Rejection Reason'] = reason_col
    
    return df_subset, rejected_df


def apply_business_rules(df, mapped_columns, all_columns=None):
    """
    Apply business validation rules to cleaned data using vectorized pandas operations.
    
    Returns: (valid_df, rejected_df)
    
    Rejected rows include a 'Rejection Reason' column.
    """
    if all_columns is None:
        all_columns = mapped_columns

    has_outstanding = 'CurrentBalanceAmt' in all_columns
    
    if not has_outstanding:
        return df, pd.DataFrame()

    # Convert CurrentBalanceAmt to numeric; non-numeric → NaN
    balance = pd.to_numeric(df['CurrentBalanceAmt'], errors='coerce')
    
    # ── Rejection masks ──
    # Rule 1: balance is null/empty
    original_empty = df['CurrentBalanceAmt'].isna() | (df['CurrentBalanceAmt'].astype(str).str.strip() == '')
    not_numeric = balance.isna() & ~original_empty
    reject_empty = original_empty
    reject_nan = not_numeric
    
    # Build helper empty-masks for each field
    def _is_empty(series):
        s = series.astype(str).str.strip().str.lower()
        return series.isna() | (s == '') | (s == 'none') | (s == 'nan')
    
    overdue_empty = _is_empty(df['overdue_amount'])
    days_empty = _is_empty(df['months_in_arrears'])
    classification_empty = _is_empty(df['loan_classification'])
    status_empty = _is_empty(df['account_status_code'])
    
    days_numeric = pd.to_numeric(df['months_in_arrears'], errors='coerce')

    # ── Account number empty check (only when account_number was mapped) ──
    account_empty = _is_empty(df['account_number']) if 'account_number' in mapped_columns else pd.Series(False, index=df.index)

    # ── Cross-field rule: "Closed"/"Cloed" in loan_classification → remap ──
    lc_lower = df['loan_classification'].astype(str).str.lower().str.strip()
    closed_mask = lc_lower.isin(['closed', 'cloed']) & ~classification_empty
    if closed_mask.any():
        df.loc[closed_mask, 'loan_classification'] = 'Performing'
        df.loc[closed_mask, 'account_status_code'] = 'Closed'
        classification_empty = classification_empty & ~closed_mask
        status_empty = status_empty & ~closed_mask
        logger.info(f"Auto-corrected {closed_mask.sum()} rows: loan_classification 'Closed'/'Cloed' -> 'Performing'")

    # ── Rule 2a: balance > 0 and months_in_arrears == 0 → auto-fill overdue = 0 ──
    bal_positive = balance > 0
    r2a_mask = bal_positive & (days_numeric == 0) & overdue_empty
    if r2a_mask.any():
        df.loc[r2a_mask, 'overdue_amount'] = 0
        overdue_empty = overdue_empty & ~r2a_mask
        logger.info(f"Auto-filled overdue_amount=0 for {r2a_mask.sum()} rows (balance>0, months_in_arrears=0)")

    # ── Rule 2a-2: balance > 0, overdue == 0, months_in_arrears missing → fill months = 0 ──
    # (If overdue > 0 and months is missing, the row will be rejected by Rule 2 below)
    overdue_numeric_r2 = pd.to_numeric(df['overdue_amount'], errors='coerce')
    r2a2_mask = bal_positive & days_empty & (overdue_numeric_r2 == 0)
    if r2a2_mask.any():
        df.loc[r2a2_mask, 'months_in_arrears'] = 0
        days_empty = days_empty & ~r2a2_mask
        days_numeric = pd.to_numeric(df['months_in_arrears'], errors='coerce')
        logger.info(f"Auto-filled months_in_arrears=0 for {r2a2_mask.sum()} rows (balance>0, overdue=0, months missing)")

    # ── Rule 2b: balance > 0, months_in_arrears == 0, overdue == 0 → force Performing + Open ──
    # Re-read overdue after Rule 2a/2a-2 may have just auto-filled zeros
    overdue_numeric = pd.to_numeric(df['overdue_amount'], errors='coerce')
    performing_open_mask = bal_positive & (days_numeric == 0) & (overdue_numeric == 0)
    if performing_open_mask.any():
        df.loc[performing_open_mask, 'loan_classification'] = 'Performing'
        df.loc[performing_open_mask, 'account_status_code'] = 'Open'
        # Update empty masks so rejection check doesn't flag these rows
        classification_empty = classification_empty & ~performing_open_mask
        status_empty = status_empty & ~performing_open_mask
        logger.info(
            f"Auto-corrected {performing_open_mask.sum()} rows: "
            f"balance>0, months_in_arrears=0, overdue=0 → loan_classification=Performing, account_status_code=Open"
        )

    # ── Rule 2: balance > 0 but critical fields still empty → reject ──
    missing_overdue = bal_positive & overdue_empty
    missing_days = bal_positive & days_empty
    missing_class = bal_positive & classification_empty
    reject_bal_positive = missing_overdue | missing_days | missing_class

    # ── Rule 3: balance == 0 → auto-fill missing fields ──
    bal_zero = balance == 0
    
    fill_overdue = bal_zero & overdue_empty
    fill_days = bal_zero & days_empty
    fill_class = bal_zero & classification_empty
    fill_status = bal_zero & status_empty

    # Rule 3b: balance=0 and months=0 given → ensure overdue=0
    fill_overdue_3b = bal_zero & (days_numeric == 0) & overdue_empty
    fill_overdue = fill_overdue | fill_overdue_3b

    if fill_overdue.any():
        df.loc[fill_overdue, 'overdue_amount'] = 0
        logger.info(f"Auto-filled overdue_amount=0 for {fill_overdue.sum()} rows (balance=0)")
    if fill_days.any():
        df.loc[fill_days, 'months_in_arrears'] = 0
        logger.info(f"Auto-filled months_in_arrears=0 for {fill_days.sum()} rows (balance=0)")
    if fill_class.any():
        df.loc[fill_class, 'loan_classification'] = 'Performing'
        logger.info(f"Auto-filled loan_classification=Performing for {fill_class.sum()} rows (balance=0)")
    if fill_status.any():
        df.loc[fill_status, 'account_status_code'] = 'Closed'
        logger.info(f"Auto-filled account_status_code=Closed for {fill_status.sum()} rows (balance=0)")

    # ── Rule 3c: balance=0, overdue=0, months_in_arrears > 0 → correct months to 0 ──
    overdue_after3 = pd.to_numeric(df['overdue_amount'], errors='coerce')
    days_after3 = pd.to_numeric(df['months_in_arrears'], errors='coerce')
    r3c_mask = bal_zero & (overdue_after3 == 0) & (days_after3 > 0)
    if r3c_mask.any():
        df.loc[r3c_mask, 'months_in_arrears'] = 0
        logger.info(f"Auto-corrected months_in_arrears to 0 for {r3c_mask.sum()} rows (balance=0, overdue=0, months>0)")

    # ── Rule 4: balance is zero or missing but overdue_amount > 0 → reject ──
    # Re-read overdue after all auto-fills so Rule 3 zeros are accounted for
    overdue_final = pd.to_numeric(df['overdue_amount'], errors='coerce')
    reject_zero_bal_with_overdue = (bal_zero | original_empty | balance.isna()) & (overdue_final > 0)

    # ── Rule 5: account_number is missing → reject ──
    reject_missing_acct = account_empty

    # ── Rules 6 & 7: cross-field overdue vs balance checks (re-read after all auto-fills) ──
    overdue_final_pre = pd.to_numeric(df['overdue_amount'], errors='coerce')
    days_empty_final = _is_empty(df['months_in_arrears'])

    # Rule 6: balance == overdue (same non-zero value) and months_in_arrears missing → reject
    reject_equal_no_months = (
        balance.notna() & overdue_final_pre.notna() &
        (balance > 0) &
        (balance == overdue_final_pre) &
        days_empty_final
    )

    # Rule 7: overdue > balance (strictly greater, not equal) → reject
    # reject_overdue_exceeds_balance = (
    #     balance.notna() & overdue_final_pre.notna() &
    #     (overdue_final_pre > balance) &
    #     (overdue_final_pre != balance)
    # )
    reject_overdue_exceeds_balance = pd.Series(False, index=df.index)  # temporarily disabled

    # ── Build rejection reasons ──
    reasons = pd.Series('', index=df.index)
    reasons = reasons.where(~reject_empty, 'Current Balance Amount is empty')
    reasons = reasons.where(~reject_nan, 'Current Balance Amount is not a valid number')
    reasons = reasons.where(~reject_zero_bal_with_overdue, 'AmountOverdue > 0 but CurrentBalanceAmt is zero or missing')
    reasons = reasons.where(~reject_missing_acct, 'AccountNo is missing')
    reasons = reasons.where(~reject_equal_no_months, 'AmountOverdue equals CurrentBalanceAmt but MonthsInArrears is missing')
    # reasons = reasons.where(~reject_overdue_exceeds_balance, 'AmountOverdue exceeds CurrentBalanceAmt')  # Rule 7 temporarily disabled
    
    # Build "missing: x, y, z" reasons for balance > 0 rejections
    reject_bal_only = reject_bal_positive & ~reject_empty & ~reject_nan
    if reject_bal_only.any():
        parts = []
        if missing_overdue.any():
            parts.append(('overdue_amount', missing_overdue))
        if missing_days.any():
            parts.append(('months_in_arrears', missing_days))
        if missing_class.any():
            parts.append(('loan_classification', missing_class))
        
        for field, mask in parts:
            combined = reject_bal_only & mask
            reasons = reasons.where(
                ~combined,
                reasons.where(
                    reasons == '',
                    reasons + ', ' + field
                ).where(reasons != '', 'Current Balance > 0 but missing: ' + field)
            )

    # ── Split valid / rejected ──
    all_reject = (
        reject_empty | reject_nan | reject_bal_positive | reject_zero_bal_with_overdue |
        reject_missing_acct | reject_equal_no_months | reject_overdue_exceeds_balance
    )
    
    if all_reject.any():
        rejected_df = df.loc[all_reject].copy()
        rejected_df['Rejection Reason'] = reasons[all_reject]
        df = df.loc[~all_reject]
        logger.info(
            f"Rejected {all_reject.sum()} rows total ("
            f"{reject_empty.sum()} empty balance, {reject_nan.sum()} non-numeric, "
            f"{reject_bal_positive.sum()} missing fields, {reject_zero_bal_with_overdue.sum()} zero-balance with overdue, "
            f"{reject_missing_acct.sum()} missing account, {reject_equal_no_months.sum()} equal balance/overdue no months, "
            f"{reject_overdue_exceeds_balance.sum()} overdue>balance)"
        )
    else:
        rejected_df = pd.DataFrame()
    
    return df, rejected_df


EXCEL_MAX_ROWS = 1_048_576
EXCEL_MAX_DATA_ROWS_PER_SHEET = EXCEL_MAX_ROWS - 1  # header uses first row


def to_excel_safe_sheet_name(name):
    """Excel sheet names cannot contain []:*?/\\ and must be <=31 chars."""
    safe = re.sub(r'[\[\]\:\*\?/\\]', '_', str(name)).strip()
    if not safe:
        safe = 'cleaned'
    return safe[:31]


def validate_excel_workbook(file_path):
    """Check if an .xlsx file is a valid zip-based workbook."""
    if not os.path.exists(file_path):
        return False
    if not zipfile.is_zipfile(file_path):
        return False
    try:
        check_wb = load_workbook(file_path, read_only=True)
        check_wb.close()
        return True
    except Exception:
        return False


def save_workbook_atomically(wb, workbook_path):
    """Save workbook via temp file → validate → replace original. Prevents corruption."""
    temp_path = f"{workbook_path}.tmp.xlsx"
    wb.save(temp_path)
    if not validate_excel_workbook(temp_path):
        raise OSError(f"Temporary workbook is invalid: {temp_path}")
    os.replace(temp_path, workbook_path)


def format_excel_sheet(ws, display_headers=None):
    """
    Apply formatting to a worksheet in a single row-iteration pass.

    Faster than per-column loops: builds a col_index→format_type map once
    from the header row, then uses ws.iter_rows() (which yields cells
    directly) rather than repeated ws.cell(row, col) dict lookups.

    Formats applied:
    - Numeric columns: 'General' number format
    - Text columns:    '@' text format + values rewritten as str
    - AccountNo:       '@' text format + left-aligned
    - Header row:      bold and border stripped
    """
    header_row_idx = 1

    numeric_names = frozenset([
        'CurrentBalanceAmt', 'Current Balance Amount',
        'overdue_amount', 'Overdue Amount', 'AmountOverdue',
        'months_in_arrears', 'Months In Arrears', 'MonthsInArrears',
        'Monthsinarrears',
    ])
    text_names = frozenset([
        'account_number', 'Account Number', 'AccountNo',
        'account_status_code', 'Account Status Code', 'AccountStatusCode',
        'loan_classification', 'Loan Classification', 'LoanClassification',
    ])
    account_no_names = frozenset(['account_number', 'Account Number', 'AccountNo'])

    # Build col_idx → format_type dict and strip header styling in one header pass
    col_format = {}  # col_idx (1-based) -> 'numeric' | 'text' | 'account'
    for cell in ws[header_row_idx]:
        name = cell.value
        if name in account_no_names:
            col_format[cell.column] = 'account'
        elif name in text_names:
            col_format[cell.column] = 'text'
        elif name in numeric_names:
            col_format[cell.column] = 'numeric'
        # Strip bold / border from every header cell
        cell.font = Font(
            name=cell.font.name, sz=cell.font.sz, bold=False,
            italic=cell.font.italic, underline=None, color=cell.font.color,
        )
        cell.border = Border()

    if not col_format or ws.max_row < 2:
        return

    # Create shared style objects once — reused for every matching cell
    left_align = Alignment(horizontal='left', vertical='center')

    # Single pass over all data rows; iter_rows yields cells directly (no dict lookup)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            fmt = col_format.get(cell.column)
            if fmt == 'numeric':
                cell.number_format = 'General'
            elif fmt in ('text', 'account'):
                cell.number_format = '@'
                if cell.value is not None:
                    cell.value = str(cell.value)
                if fmt == 'account':
                    cell.alignment = left_align


def save_df_to_excel_robust(df, file_path, sheet_name='cleaned', chunk_size=None):
    """
    Save DataFrame to Excel with:
    - Multi-sheet splitting if data exceeds Excel row limit
    - Atomic save (write to temp, validate, replace)
    - Corruption recovery (backup corrupt files)
    - PermissionError fallback chain (3 attempts)
    - Per-sheet formatting
    
    Returns: (final_path, sheet_names_list)
    """
    if chunk_size is None:
        chunk_size = EXCEL_MAX_DATA_ROWS_PER_SHEET

    sheet_name = to_excel_safe_sheet_name(sheet_name)

    def _build_workbook(dataframe, base_sheet_name):
        """Build an openpyxl Workbook from DataFrame, splitting if needed."""
        wb = Workbook()
        # Remove default sheet
        if wb.active:
            wb.remove(wb.active)

        total_rows = len(dataframe)
        num_parts = max(1, int(np.ceil(total_rows / chunk_size)))
        saved_names = []

        for part_idx in range(num_parts):
            start = part_idx * chunk_size
            end = (part_idx + 1) * chunk_size
            chunk_df = dataframe.iloc[start:end]

            if num_parts == 1:
                part_name = base_sheet_name
            else:
                part_name = to_excel_safe_sheet_name(f"{base_sheet_name}_part{part_idx + 1}")

            ws = wb.create_sheet(title=part_name)
            saved_names.append(part_name)

            for row in dataframe_to_rows(chunk_df, index=False, header=True):
                ws.append(row)

            format_excel_sheet(ws)

        return wb, saved_names

    def _try_save(target_path):
        """Attempt to build and atomically save the workbook."""
        # If file exists but is corrupt, back it up
        if os.path.exists(target_path):
            if not validate_excel_workbook(target_path):
                corrupted_backup = f"{target_path}.corrupt_{datetime.now().strftime('%H%M%S')}"
                try:
                    os.replace(target_path, corrupted_backup)
                    logger.warning(f"Backed up corrupt workbook: {corrupted_backup}")
                except OSError:
                    pass

        wb, names = _build_workbook(df, sheet_name)
        save_workbook_atomically(wb, target_path)
        return target_path, names

    # Attempt 1: primary path
    try:
        return _try_save(file_path)
    except PermissionError:
        logger.warning(f"PermissionError on {file_path}, trying timestamped fallback")

    # Attempt 2: timestamped fallback (file might be open by user)
    base, ext = os.path.splitext(file_path)
    fallback_path = f"{base}_{datetime.now().strftime('%H%M%S')}{ext}"
    try:
        return _try_save(fallback_path)
    except PermissionError:
        logger.warning(f"PermissionError on {fallback_path}, trying last resort")

    # Attempt 3: last resort with microseconds
    last_resort_path = f"{base}_recover_{datetime.now().strftime('%H%M%S%f')}{ext}"
    return _try_save(last_resort_path)


def get_excel_sheet_names(file_path):
    """Return list of sheet names from an Excel file. Returns None for CSV."""
    ext = os.path.splitext(file_path)[1].lower()
    if ext == '.csv':
        return None
    try:
        xl = pd.ExcelFile(file_path)
        return xl.sheet_names
    except Exception:
        return None


def read_uploaded_file_sheet(file_path, sheet_name=None, header=0):
    """Read a specific sheet from an Excel file, or the full CSV."""
    ext = os.path.splitext(file_path)[1].lower()
    with warnings.catch_warnings():
        warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
        if ext == '.csv':
            return pd.read_csv(file_path, dtype=str, header=header)
        elif ext == '.xlsx':
            return pd.read_excel(file_path, sheet_name=sheet_name, dtype=str, engine='openpyxl', header=header)
        elif ext == '.xls':
            return pd.read_excel(file_path, sheet_name=sheet_name, dtype=str, engine='xlrd', header=header)
        else:
            try:
                return pd.read_excel(file_path, sheet_name=sheet_name, dtype=str, engine='openpyxl', header=header)
            except Exception:
                return pd.read_excel(file_path, sheet_name=sheet_name, dtype=str, engine='xlrd', header=header)


def excel_to_csv_streaming(file_path, output_path, sheet_name=None, header_row=0):
    """
    Convert an .xlsx file to CSV using openpyxl read_only streaming mode.

    Reads row-by-row without loading the full workbook into RAM, making it
    safe for files that would otherwise exhaust memory via pd.read_excel().

    Args:
        file_path:   source .xlsx file path
        output_path: destination .csv path (will be overwritten if it exists)
        sheet_name:  sheet to read; None = first/active sheet
        header_row:  0-based row index of the header; rows before it are skipped

    Returns: output_path

    Note: only .xlsx is supported (openpyxl read_only). Callers must ensure
    they only call this for .xlsx files; .xls falls back to full pd.read_excel().
    """
    import csv as _csv

    with warnings.catch_warnings():
        warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
        wb = load_workbook(file_path, read_only=True, data_only=True)
        try:
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.active

            with open(output_path, 'w', newline='', encoding='utf-8') as f:
                writer = _csv.writer(f)
                for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                    if row_idx < header_row:
                        continue  # skip pre-header rows
                    writer.writerow(['' if v is None else str(v) for v in row])
        finally:
            wb.close()

    logger.info(f"Stream-converted Excel to CSV: {file_path} → {output_path} (skipped {header_row} pre-header rows)")
    return output_path


def extract_sub_id(sheet_name):
    """Extract subscriber ID from sheet name pattern: subid_date_client."""
    parts = str(sheet_name).split('_')
    if parts:
        return parts[0]
    return ''


def build_sheet_name(subscriber, date=None, index=None):
    """
    Build a normalised sheet name from a Subscriber instance.
    Pattern: {subscriber_id}_{ddmmyyyy}_{subscriber_name_lower}
    Appends _{index} suffix for multi-sheet uploads (index >= 2).

    Args:
        subscriber: Subscriber model instance
        date: datetime.date or datetime.datetime; defaults to today
        index: int or None — if provided and >= 2, appended as suffix
    Returns:
        str sheet name safe for use as Excel sheet name and SQL table name
    """
    from datetime import date as _date
    d = date or _date.today()
    base = f"{subscriber.subscriber_id}_{d.strftime('%d%m%Y')}_{subscriber.subscriber_name.lower()}"
    if index is not None and index >= 2:
        base = f"{base}_{index}"
    return base


def generate_sql_script(sheet_name):
    """
    Generate a SQL script by replacing the template base name and subscriber ID
    with values derived from the sheet name.
    Returns the path to the generated script file.
    """
    template_path = settings.SQL_TEMPLATE_PATH
    base_name = settings.SQL_TEMPLATE_BASE_NAME
    base_sub_id = settings.SQL_TEMPLATE_BASE_SUBID

    with open(template_path, 'r', encoding='utf-8') as f:
        script = f.read()

    # Replace all instances of the base table name (handles V1, V2, V1_comm, etc.)
    script = script.replace(base_name, sheet_name)

    # Replace subscriber ID
    new_sub_id = extract_sub_id(sheet_name)
    if new_sub_id and new_sub_id != base_sub_id:
        script = script.replace(f'SubscriberID={base_sub_id}', f'SubscriberID={new_sub_id}')

    # Save generated script
    os.makedirs(settings.GENERATED_SCRIPTS_DIR, exist_ok=True)
    output_path = os.path.join(settings.GENERATED_SCRIPTS_DIR, f'{sheet_name}.sql')
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(script)

    logger.info(f"Generated SQL script for sheet '{sheet_name}': {output_path}")
    return output_path


def get_batchupdate_connection():
    """Get a pyodbc connection to the BatchUpdate database."""
    conn_str = (
        f"DRIVER={{{settings.BATCHUPDATE_DRIVER}}};"
        f"SERVER={settings.BATCHUPDATE_SERVER};"
        f"DATABASE={settings.BATCHUPDATE_DB};"
    )
    if settings.BATCHUPDATE_TRUSTED_CONNECTION.lower() == 'yes':
        conn_str += "Trusted_Connection=yes;"
    return pyodbc.connect(conn_str)


def get_subscribers_from_batchupdate():
    """
    Return subscriber list from BatchUpdate's Sheet1 table.
    Result is cached for SUBSCRIBER_CACHE_TTL seconds (default 5 min) to avoid
    a SQL Server round-trip on every page load.
    Returns a list of dicts: [{'subscriber_id': int, 'subscriber_name': str}, ...]
    """
    from django.core.cache import cache
    from django.conf import settings as _settings
    from .models import BatchSubscriber

    cache_key = 'subscribers_list'
    cached = cache.get(cache_key)
    if cached is not None:
        return cached

    result = [
        {'subscriber_id': int(row['subscriber_id']), 'subscriber_name': row['subscriber_name']}
        for row in BatchSubscriber.objects.order_by('subscriber_name').values('subscriber_id', 'subscriber_name')
    ]
    ttl = getattr(_settings, 'SUBSCRIBER_CACHE_TTL', 300)
    cache.set(cache_key, result, ttl)
    return result


def upload_raw_to_batchupdate(df, table_name):
    """
    Upload a raw DataFrame to the BatchUpdate database as a new table.
    Table name = sheet name. Columns come from DataFrame headers.
    Drops existing table if present, then creates and bulk-inserts.
    """
    # SQL types for known display-header column names — defined in columns.py
    from .columns import COLUMN_SQL_TYPES

    conn = get_batchupdate_connection()
    cursor = conn.cursor()
    cursor.fast_executemany = True

    # Escape ] as ]] to prevent bracket-identifier injection (both table name and column names)
    safe_table = table_name.replace("]", "]]").replace("'", "''")

    try:
        # Drop table if exists
        cursor.execute(f"IF OBJECT_ID('[{safe_table}]', 'U') IS NOT NULL DROP TABLE [{safe_table}]")
        conn.commit()

        # Build CREATE TABLE using proper types for known columns, NVARCHAR(255) for others
        col_defs = ', '.join(
            f'[{col.replace("]", "]]")}] {COLUMN_SQL_TYPES.get(col, "NVARCHAR(255)")}'
            for col in df.columns
        )
        cursor.execute(f"CREATE TABLE [{safe_table}] ({col_defs})")
        conn.commit()

        # Bulk insert using executemany with fast_executemany enabled
        placeholders = ', '.join(['?'] * len(df.columns))
        insert_sql = f"INSERT INTO [{safe_table}] VALUES ({placeholders})"

        # Convert NaN to None for SQL NULL
        data = df.where(df.notna(), None).values.tolist()

        batch_size = 5000
        for i in range(0, len(data), batch_size):
            batch = data[i:i + batch_size]
            cursor.executemany(insert_sql, batch)
            conn.commit()

        return len(data)

    except Exception as e:
        conn.rollback()
        logger.error(f"BatchUpdate upload failed for [{safe_table}]: {e}", exc_info=True)
        raise
    finally:
        cursor.close()
        conn.close()


def should_use_chunking(file_path):
    """
    Decide whether to use chunked processing based on file type and disk size.
    - Excel: Chunk if size > 15MB (~50,000 rows in Openpyxl memory model)
    - CSV: Chunk if size > 30MB (~200,000 rows in plain CSV)
    """
    ext = os.path.splitext(file_path)[1].lower()
    try:
        file_size_mb = os.path.getsize(file_path) / (1024 * 1024)
    except OSError:
        return False

    if ext in ('.xlsx', '.xls'):
        return file_size_mb > 15
    else:
        return file_size_mb > 30


def upload_parquet_to_batchupdate(parquet_path, table_name):
    """
    Read a Parquet file in chunks and stream-upload it to the BatchUpdate database.
    Prevents memory spikes by iterating over Arrow record batches.
    """
    import pyarrow.parquet as pq
    from .columns import COLUMN_SQL_TYPES

    conn = get_batchupdate_connection()
    cursor = conn.cursor()
    cursor.fast_executemany = True

    # Escape ] as ]] to prevent bracket-identifier injection (both table name and column names)
    safe_table = table_name.replace("]", "]]").replace("'", "''")

    try:
        # Drop table if exists
        cursor.execute(f"IF OBJECT_ID('[{safe_table}]', 'U') IS NOT NULL DROP TABLE [{safe_table}]")
        conn.commit()

        # Open the Parquet file and read headers from the first batch
        parquet_file = pq.ParquetFile(parquet_path)
        first_batch = next(parquet_file.iter_batches(batch_size=1))
        columns = first_batch.schema.names

        # Build CREATE TABLE using proper types for known columns, NVARCHAR(255) for others
        col_defs = ', '.join(
            f'[{col.replace("]", "]]")}] {COLUMN_SQL_TYPES.get(col, "NVARCHAR(255)")}'
            for col in columns
        )
        cursor.execute(f"CREATE TABLE [{safe_table}] ({col_defs})")
        conn.commit()

        # Bulk insert using executemany with fast_executemany enabled
        placeholders = ', '.join(['?'] * len(columns))
        insert_sql = f"INSERT INTO [{safe_table}] VALUES ({placeholders})"

        # Stream batches into the database
        for batch in parquet_file.iter_batches(batch_size=5000):
            df = batch.to_pandas()
            # Convert NaN to None for SQL NULL
            data = df.where(df.notna(), None).values.tolist()
            cursor.executemany(insert_sql, data)
            conn.commit()

        logger.info(f"Stream-uploaded Parquet to BatchUpdate table [{safe_table}]")
        return parquet_file.metadata.num_rows

    except Exception as e:
        conn.rollback()
        logger.error(f"Parquet BatchUpdate upload failed for [{safe_table}]: {e}", exc_info=True)
        raise
    finally:
        cursor.close()
        conn.close()

